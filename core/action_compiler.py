import logging
import os
from typing import Any, Dict, List, Union

from jinja2 import BaseLoader, Environment
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from action_times import (
    DOG_DEFAULT_ACTION_TIMES,
    DOG_PATTERN_FALLBACK_TIMES,
    DRONE_DEFAULT_ACTION_TIMES,
    DRONE_PATTERN_FALLBACK_TIMES,
)
from core.spreadsheet_loader import SpreadsheetLoader


class ActionCompiler:
    """
    Compiles and validates robot action sequences from spreadsheet data.

    This class is responsible for:
    1. Compiling action sequences from spreadsheet data
    2. Validating that all actions exist in the action details
    3. Ensuring that action execution times don't exceed allotted time slots
    """

    def __init__(self, spreadsheet_loader: SpreadsheetLoader):
        """
        Initialize the ActionCompiler with a SpreadsheetLoader.

        Args:
            spreadsheet_loader: A loader that provides access to spreadsheet data
        """
        self.spreadsheet_loader = spreadsheet_loader
        self.logger = logging.getLogger("ActionCompiler")

    def _get_robot_keys(self, action: Dict[str, str]) -> List[str]:
        """Helper to get all robot keys in an action row.

        Returns keys that start with:
        - Humanoid_ (humanoid robots)
        - Drone_ (drone robots)
        - Dog_ (dog robots)
        """
        return [
            key for key in action if key.startswith(("Humanoid_", "Drone_", "Dog_"))
        ]

    def _get_dog_default_actions(self) -> Dict[str, float]:
        """Extract default dog action timings from constants."""
        return DOG_DEFAULT_ACTION_TIMES.copy()

    def _get_drone_default_actions(self) -> Dict[str, float]:
        """Extract default drone action timings from constants."""
        return DRONE_DEFAULT_ACTION_TIMES.copy()

    def _get_enhanced_action_name_to_time(self) -> Dict[str, float]:
        """Get enhanced action name to time mapping that includes default dog and drone actions."""
        # Start with spreadsheet data and ensure all values are floats
        base_actions = self.spreadsheet_loader.get_action_name_to_time()
        action_name_to_time = {}

        # Convert all values to floats
        for name, time_val in base_actions.items():
            try:
                action_name_to_time[name] = float(time_val)
            except (ValueError, TypeError):
                self.logger.warning(
                    f"Invalid time value for action '{name}': {time_val}"
                )
                continue

        # Add dog default actions
        dog_actions = self._get_dog_default_actions()
        for action, time_val in dog_actions.items():
            if action not in action_name_to_time:
                action_name_to_time[action] = time_val

        # Add drone default actions
        drone_actions = self._get_drone_default_actions()
        for action, time_val in drone_actions.items():
            if action not in action_name_to_time:
                action_name_to_time[action] = time_val

        self.logger.debug(
            f"Enhanced action details loaded: {len(action_name_to_time)} actions"
        )
        return action_name_to_time

    def _get_action_time_with_fallback(
        self, action_name: str, action_name_to_time: Dict[str, Union[str, float]]
    ) -> float:
        """
        Get action time with fallback logic for parameterized drone and dog commands.

        Args:
            action_name: The action name to look up
            action_name_to_time: The action mapping dictionary

        Returns:
            The action time, or 0 if not found
        """
        action_lower = action_name.lower()

        # First try exact match (original case and lowercase)
        if action_name in action_name_to_time:
            value = action_name_to_time[action_name]
            return float(value) if isinstance(value, (str, int, float)) else 0.0
        if action_lower in action_name_to_time:
            value = action_name_to_time[action_lower]
            return float(value) if isinstance(value, (str, int, float)) else 0.0

        # Enhanced drone action pattern matching
        if action_lower.startswith(("move_", "rotate_", "flip_")):
            # Extract base command (e.g., "move_up_100" -> "move_up")
            parts = action_lower.split("_")
            if len(parts) >= 2:
                base_command = "_".join(parts[:2])  # e.g., "move_up"
                if base_command in action_name_to_time:
                    return float(action_name_to_time[base_command])

        # Try more specific drone action patterns (ordered by specificity)
        for pattern, default_time in DRONE_PATTERN_FALLBACK_TIMES.items():
            if action_lower.startswith(pattern):
                self.logger.debug(
                    f"Action '{action_name}' matched drone pattern '{pattern}' with time {default_time}s"
                )
                return default_time

        # Enhanced dog action pattern matching (check exact match first, then prefix)
        for pattern, default_time in DOG_PATTERN_FALLBACK_TIMES.items():
            if action_lower == pattern or action_lower.startswith(pattern + "_"):
                self.logger.debug(
                    f"Action '{action_name}' matched dog pattern '{pattern}' with time {default_time}s"
                )
                return default_time

        # Additional pattern matching for common action variations
        # Handle parameterized actions that might have numbers or additional suffixes
        for base_action in action_name_to_time:
            if isinstance(base_action, str) and action_lower.startswith(
                base_action.lower() + "_"
            ):
                # Found a parameterized version of a known action
                value = action_name_to_time[base_action]
                if isinstance(value, (str, int, float)):
                    self.logger.debug(
                        f"Action '{action_name}' matched base action '{base_action}' "
                        f"with time {float(value)}s"
                    )
                    return float(value)

        return 0.0  # Not found

    def _expand_macro_aliases(self, value: str) -> str:
        """
        Expand macro aliases in the action value with their corresponding action sequences.

        Args:
            value: The action value that may contain macro aliases

        Returns:
            The expanded action value with aliases replaced by their action sequences
        """
        if not value:
            return value

        # Get macro aliases mapping
        macro_aliases = self.spreadsheet_loader.get_macro_aliases()
        if not macro_aliases:
            return value

        # Split the value into lines to process each action
        lines = value.splitlines()
        expanded_lines = []

        for line in lines:
            line = line.strip()
            if not line:
                continue

            # Check if this line is a macro alias
            if line in macro_aliases:
                # Replace with the macro's action sequence
                macro_actions = macro_aliases[line]
                # Split macro actions by newlines and add each as separate lines
                macro_lines = [
                    ml.strip() for ml in macro_actions.splitlines() if ml.strip()
                ]
                expanded_lines.extend(macro_lines)
                self.logger.debug(
                    f"Expanded macro alias '{line}' to {len(macro_lines)} actions"
                )
            else:
                # Keep the original line
                expanded_lines.append(line)

        return "\n".join(expanded_lines)

    def compile_actions(
        self, export_excel: bool = False, csv_path: str = None, song_name: str = None
    ) -> List[Dict[str, Any]]:
        """
        Compile and validate robot actions from spreadsheet data.

        Args:
            export_csv: If True, also export the compiled actions to CSV with duration annotations
            csv_path: Optional path for the CSV export. If None, saves to data folder
            song_name: Optional song name for better cache filename generation

        Returns:
            List of dictionaries containing validated robot actions

        Raises:
            ValueError: If actions don't exist or exceed their time allocation
        """
        robot_actions = self.spreadsheet_loader.get_robot_actions()
        action_name_to_time = self._get_enhanced_action_name_to_time()

        for action in robot_actions:
            for key in self._get_robot_keys(action):
                value = action[key]

                # First expand any macro aliases to their action sequences
                if value:
                    value = self._expand_macro_aliases(value)

                # Then render as Jinja2 template if there are template markers
                if value and ("{{" in value or "}}" in value):
                    rtemplate = Environment(loader=BaseLoader).from_string(value)
                    value = rtemplate.render({})

                # Update the action with the processed value
                action[key] = value

        self.logger.info(f"Compiled {len(robot_actions)} action sequences")
        self.logger.debug(f"Action details loaded: {list(action_name_to_time.keys())}")

        self.check_actions_existence(
            robot_actions, action_name_to_time, strict_mode=False
        )
        self.check_actions_time(robot_actions, action_name_to_time)

        # Optional Excel export to data folder
        if export_excel:
            excel_file = self.export_compiled_actions_to_excel(csv_path, song_name)
            self.logger.info(f"Actions exported to Excel cache: {excel_file}")

        return robot_actions

    def export_compiled_actions_to_excel(
        self, output_path: str = None, song_name: str = None
    ) -> str:
        """
        Export compiled actions to Excel format with duration annotations and formatting.
        Files are saved to the data folder as cache.

        Each action in robot columns will be formatted as: action_name (duration)
        Multiline actions are preserved in the Excel cells with proper formatting.

        Args:
            output_path: Optional path to save the Excel file. If None, auto-generates in data folder
            song_name: Optional song name for cache-friendly filename generation

        Returns:
            The path to the exported Excel file
        """
        # Auto-generate path if not provided
        if output_path is None:
            # Create data folder if it doesn't exist
            data_folder = os.path.join(
                os.path.dirname(os.path.dirname(__file__)), "data"
            )
            os.makedirs(data_folder, exist_ok=True)

            # Generate cache-friendly filename
            if song_name:
                # Use song name for filename
                safe_song_name = "".join(
                    c for c in song_name if c.isalnum() or c in ("-", "_")
                )
                filename = f"compiled_actions_{safe_song_name}.xlsx"
            else:
                # Fallback to spreadsheet ID
                spreadsheet_id = getattr(
                    self.spreadsheet_loader, "spreadsheet_id", "unknown"
                )
                safe_id = "".join(
                    c for c in str(spreadsheet_id) if c.isalnum() or c in ("-", "_")
                )
                filename = f"compiled_actions_{safe_id}.xlsx"

            output_path = os.path.join(data_folder, filename)

        # Ensure output directory exists
        output_dir = os.path.dirname(output_path)
        if output_dir:
            os.makedirs(output_dir, exist_ok=True)

        robot_actions = self.spreadsheet_loader.get_robot_actions()
        action_name_to_time = self._get_enhanced_action_name_to_time()

        # Process actions to add duration annotations
        processed_actions = []
        for action in robot_actions:
            processed_action = action.copy()

            for key in self._get_robot_keys(action):
                value = action[key]
                if value:
                    # Handle Jinja2 templates
                    if "{{" in value or "}}" in value:
                        rtemplate = Environment(loader=BaseLoader).from_string(value)
                        value = rtemplate.render({})

                    # Process each action line
                    action_lines = [a.strip() for a in value.splitlines() if a.strip()]
                    annotated_lines = []
                    total_cell_time = 0.0

                    for act in action_lines:
                        act_time = self._get_action_time_with_fallback(
                            act, action_name_to_time
                        )
                        total_cell_time += act_time
                        if act_time > 0:
                            annotated_lines.append(f"{act} ({act_time} s)")
                        else:
                            annotated_lines.append(f"{act} (0 s)")

                    # Add total time at the BEGINNING if there are actions
                    if annotated_lines:
                        # Insert total time as first line with clear formatting
                        annotated_lines.insert(0, f"⏱️ TOTAL: {total_cell_time} s")

                    # Join back with newlines to preserve multiline format
                    processed_action[key] = "\n".join(annotated_lines)

            processed_actions.append(processed_action)

        # Create Excel workbook with formatting
        if processed_actions:
            wb = Workbook()
            ws = wb.active
            ws.title = "Robot Actions"

            # Get column headers
            headers = list(processed_actions[0].keys())

            # Style definitions
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            time_fill = PatternFill(
                start_color="E6F3FF", end_color="E6F3FF", fill_type="solid"
            )
            robot_fill = PatternFill(
                start_color="F0F8FF", end_color="F0F8FF", fill_type="solid"
            )
            total_time_font = Font(bold=True, color="0066CC", size=11)

            # Write headers with formatting
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Write data rows
            for row_idx, action in enumerate(processed_actions, 2):
                for col_idx, header in enumerate(headers, 1):
                    cell_value = action.get(header, "")
                    cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)

                    # Apply formatting based on column type
                    if header == "Time":
                        cell.fill = time_fill
                        cell.alignment = Alignment(
                            horizontal="center", vertical="center"
                        )
                        cell.font = Font(bold=True)
                    elif header.startswith(("Humanoid_", "Drone_", "Dog_")):
                        cell.fill = robot_fill
                        cell.alignment = Alignment(
                            horizontal="left", vertical="top", wrap_text=True
                        )

                        # Check if cell starts with total time and apply special formatting
                        if cell_value and cell_value.startswith("⏱️ TOTAL:"):
                            # Apply special formatting for cells with totals at the beginning
                            cell.font = total_time_font
                        else:
                            cell.font = Font(size=10)
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")

            # Auto-adjust column widths
            for col_idx, header in enumerate(headers, 1):
                column_letter = get_column_letter(col_idx)
                if header == "Time":
                    ws.column_dimensions[column_letter].width = 8
                elif header.startswith(("Humanoid_", "Drone_", "Dog_")):
                    ws.column_dimensions[column_letter].width = 25
                else:
                    ws.column_dimensions[column_letter].width = 15

            # Adjust row heights for better readability
            for row_idx in range(2, len(processed_actions) + 2):
                ws.row_dimensions[row_idx].height = None  # Auto height
                # Check if any cell in this row has multiple lines
                for col_idx, header in enumerate(headers, 1):
                    if header.startswith(("Humanoid_", "Drone_", "Dog_")):
                        cell_value = processed_actions[row_idx - 2].get(header, "")
                        if cell_value and "\n" in cell_value:
                            # Set minimum height for multiline cells
                            line_count = cell_value.count("\n") + 1
                            min_height = max(20, line_count * 15)
                            ws.row_dimensions[row_idx].height = min_height
                            break

            # Save the workbook
            wb.save(output_path)

        self.logger.info(
            f"Exported {len(processed_actions)} compiled actions to {output_path}"
        )
        return output_path

    # Keep the old CSV method for backward compatibility
    def export_compiled_actions_to_csv(
        self, output_path: str = None, song_name: str = None
    ) -> str:
        """
        Legacy CSV export method - redirects to Excel export.
        Maintained for backward compatibility.
        """
        return self.export_compiled_actions_to_excel(output_path, song_name)

    def check_actions_time(
        self,
        robot_actions: List[Dict[str, str]],
        action_name_to_time: Dict[str, Union[str, float]],
    ) -> None:
        """
        Validate that action execution times don't exceed their allocated time slot.
        Uses column names to determine robot types for better error reporting.

        Args:
            robot_actions: List of robot action sequences
            action_name_to_time: Mapping of action names to execution times

        Raises:
            ValueError: If action times exceed allocated time slot
        """
        for idx, action in enumerate(robot_actions, start=1):
            time_val = action.get("Time")
            if not time_val:
                self.logger.warning(
                    f"Row {idx}: No time value specified, skipping validation"
                )
                continue

            try:
                allocated_time = float(time_val)
            except (ValueError, TypeError):
                self.logger.warning(
                    f"Row {idx}: Invalid time value '{time_val}', skipping validation"
                )
                continue

            for key in self._get_robot_keys(action):
                # Extract robot type from column name
                robot_type = key.split("_")[0].lower()

                value = action[key]
                if value:
                    actions = [a.strip() for a in value.splitlines() if a.strip()]
                    total_time = 0.0
                    action_details = []

                    for act in actions:
                        act_time = self._get_action_time_with_fallback(
                            act, action_name_to_time
                        )
                        total_time += act_time
                        action_details.append(f"{act}({act_time}s)")

                    self.logger.debug(
                        f"Row {idx} {robot_type} {key}: {', '.join(action_details)} "
                        f"= {total_time}s / {allocated_time}s"
                    )

                    if total_time > allocated_time:
                        raise ValueError(
                            f"Row {idx}: {robot_type.title()} {key} actions exceed time limit "
                            f"({total_time}s > {allocated_time}s): {', '.join(action_details)}"
                        )

    def check_actions_existence(
        self,
        robot_actions: List[Dict[str, str]],
        action_name_to_time: Dict[str, Union[str, float]],
        strict_mode: bool = True,
    ) -> None:
        """
        Validate that all specified actions exist in the action details or are known default actions.
        Uses column names to determine robot types and validate appropriate actions.

        Args:
            robot_actions: List of robot action sequences
            action_name_to_time: Mapping of action names to execution times
            strict_mode: If True, raises ValueError for unknown actions. If False, only logs warnings.

        Raises:
            ValueError: If an action is referenced but not defined in action details or defaults (strict_mode=True)
        """
        missing_actions = {}  # Track missing actions by robot type

        for idx, action in enumerate(robot_actions, start=1):
            for key in self._get_robot_keys(action):
                # Extract robot type from column name (e.g., "Drone_1" -> "drone")
                robot_type = key.split("_")[0].lower()

                value = action[key]
                if value:
                    actions = [a.strip() for a in value.splitlines() if a.strip()]
                    for act in actions:
                        action_time = self._get_action_time_with_fallback(
                            act, action_name_to_time
                        )

                        if action_time <= 0:
                            # Track missing actions by robot type
                            if robot_type not in missing_actions:
                                missing_actions[robot_type] = set()
                            missing_actions[robot_type].add(act)

                            error_msg = f"Row {idx}: {robot_type.title()} action '{act}' not found"

                            if strict_mode:
                                raise ValueError(error_msg)
                            else:
                                self.logger.warning(
                                    f"{error_msg}. May use runtime defaults."
                                )

        # Log summary of missing actions by robot type
        if missing_actions:
            for robot_type, actions in missing_actions.items():
                action_list = sorted(list(actions))
                self.logger.info(
                    f"Missing {robot_type} actions ({len(action_list)}): {', '.join(action_list)}"
                )
